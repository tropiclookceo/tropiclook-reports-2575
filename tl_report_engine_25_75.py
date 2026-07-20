#!/usr/bin/env python3
"""
tl_report_engine_25_75.py — TropicLook Owner Report Generation Engine
NEW MODEL: 25/75 AFTER CHANNEL COSTS (OTA / agent) — v1.0

Formula chain (source of truth:
01_Executive/Company_Management_System/TL_OWNER_COOPERATION_MODEL_25_75_AFTER_OTA_STANDARD_V0_1.md):

    Gross accommodation revenue
    - channel costs (OTA / agent / channel), ACTUAL amount withheld per booking
    = NET after channel                       <- distribution base
    - TropicLook share 25% of NET
    = Owner share 75% of NET
    - Owner expenses (personal stay / asset / building only)
    = Net Owner Income -> Owner Account balance -> payouts

Standard guest cleaning / laundry / chemicals / welcome packs / guest
electricity are INSIDE TropicLook economics and MUST NOT appear as owner
expenses. Resort Fee is a separate TL revenue stream and never enters this
report. Revenue recognition: CHECKOUT DATE (unchanged canon).

Reads: INPUT template 25/75 (8-sheet xlsx) -> Writes: 6-tab Owner Report.
"""

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
from collections import defaultdict
import argparse, calendar, sys, os, io, re

ENGINE_VERSION = "25/75 after OTA v1.0"

# ── BRAND ─────────────────────────────────────────────────────────────────────
NAVY     = "1F3864"
TEAL     = "1F6E6E"
GOLD     = "C9A84C"
SILVER   = "D9D9D9"
LIGHT_BG = "EBF0F7"
WHITE    = "FFFFFF"
RED_BG   = "FFC7CE"
GREEN_BG = "C6EFCE"
FONT     = "Arial"

MONTHS_RU    = {1:"Январь",2:"Февраль",3:"Март",4:"Апрель",5:"Май",6:"Июнь",
                7:"Июль",8:"Август",9:"Сентябрь",10:"Октябрь",11:"Ноябрь",12:"Декабрь"}
MONTHS_SHORT = {1:"Jan",2:"Feb",3:"Mar",4:"Apr",5:"May",6:"Jun",
                7:"Jul",8:"Aug",9:"Sep",10:"Oct",11:"Nov",12:"Dec"}

# ── OWNER-ZONE EXPENSE TAXONOMY (new model) ──────────────────────────────────
# Only owner-zone categories are allowed in Owner_Expenses.
CAT_NAMES = {
    # Personal stay of the Owner
    "OWN-CLEAN":  "Уборка при личном проживании",
    "OWN-LNDRY":  "Стирка при личном проживании",
    "OWN-ELEC":   "Электричество при личном проживании",
    # Repairs / asset care
    "MNT-REPAIR": "Ремонтные работы",
    "MNT-CALL":   "Технический вызов",
    "EMRG":       "Аварийный ремонт",
    "MNT-AC":     "Сервис кондиционеров",
    "MNT-SEPTIC": "Очистка септика",
    "PLN-DEEP":   "Генеральная уборка (плановая)",
    "PLN-DRY":    "Химчистка текстиля / мебели",
    "CLN-DEEP":   "Генеральная уборка (плановая)",   # legacy alias
    "CLN-DRY":    "Химчистка текстиля / мебели",     # legacy alias
    # Villa service contract
    "SVC-VILLA":  "Обслуживание виллы (сервисный договор)",
    "FIX-POOL":   "Обслуживание бассейна",
    "FIX-GARDEN": "Обслуживание сада",
    "FIX-Garden": "Обслуживание сада",               # legacy alias
    "FIX-MAIN":   "Техническое обслуживание",
    "MNT-MAIN":   "Регулярное техническое обслуживание",
    # Building / utilities that stay with the Owner
    "FIX-COM":    "Общие расходы здания (CAM)",
    "FIX-INET":   "Интернет",
    "UTL-WAT":    "Вода",
    "WASTE":      "Вывоз мусора",
    "FIX-PEST":   "Пест-контроль",
    "FIX-SEC":    "Охрана / CCTV",
    "FIX-INS":    "Страховка объекта",
    # Other owner-zone
    "TAXES-PRP":  "Налог на имущество",
    "FFE-EQUIP":  "Оборудование и инвентарь (по согласованию)",
    "MISC":       "Прочее",
    "ADJ":        "Корректировки",
}

BUDGET_ORDER = [
    "SVC-VILLA","FIX-POOL","FIX-GARDEN","FIX-Garden","MNT-MAIN","FIX-MAIN",
    "FIX-COM","FIX-INET","UTL-WAT","WASTE","FIX-PEST","FIX-SEC","FIX-INS",
    "OWN-CLEAN","OWN-LNDRY","OWN-ELEC",
    "PLN-DEEP","PLN-DRY","CLN-DEEP","CLN-DRY","MNT-AC","MNT-SEPTIC",
    "MNT-REPAIR","MNT-CALL","EMRG",
    "TAXES-PRP","FFE-EQUIP","MISC","ADJ",
]

# Guest-zone codes of the OLD model. In the new model these costs live inside
# TropicLook economics and must NOT be entered as owner expenses.
FORBIDDEN_GUEST_CODES = {
    "FIX-CLEAN":  "стандартная гостевая уборка — экономика TropicLook",
    "CLN-REG":    "стандартная гостевая уборка — экономика TropicLook",
    "VAR-CLEAN":  "стандартная гостевая уборка — экономика TropicLook",
    "VAR-LNDRY":  "стандартная гостевая стирка — экономика TropicLook",
    "CLN-LNDRY":  "стандартная гостевая стирка — экономика TropicLook",
    "VAR-CHEM":   "химия и расходники для уборки — экономика TropicLook",
    "CLN-CHEM":   "химия и расходники для уборки — экономика TropicLook",
    "VAR-WELC":   "приветственные пакеты — экономика TropicLook",
    "UTL-ELEC":   "гостевое электричество — экономика TropicLook; для личного проживания используйте OWN-ELEC",
    "UTL-RCHG":   "возмещение электричества гостём отменено — гостевое электричество в экономике TropicLook",
    "GUEST-SVC":  "гостевой сервис — экономика TropicLook",
}

# Channels where a channel commission is expected. If channel matches and the
# commission is missing/zero, validation raises a warning (CEO rule 2026-05-18:
# 'Direct from booking.com' is still Booking.com).
OTA_CHANNEL_MARKERS = (
    "booking", "airbnb", "agoda", "expedia", "trip.com", "ctrip",
    "hopper", "hotels", "ota", "agent",
)

REQUIRED_INPUT_SHEETS = [
    "Property_Info",
    "Reservations",
    "Owner_Expenses",
    "Owner_Payouts",
    "OPEX_Budget",
    "Prior_Period",
    "Cumulative",
    "Cash_Balance",
]

# Legacy sheet name accepted as a fallback for Owner_Expenses.
EXPENSES_SHEET_ALIASES = ("Owner_Expenses", "Expenses")

OWNER_REPORT_SHEETS = [
    "Dashboard",
    "P&L Monthly",
    "OPEX Passport",
    "12-Month Summary",
    "Transaction Ledger",
    "DAP Snapshot",
]

PAYIN_MARKERS = ("payin", "pay-in", "пополнен", "взнос")


# ── STYLE HELPERS ─────────────────────────────────────────────────────────────
def _f(bold=False, size=10, color=None, name=FONT):
    return Font(name=name, bold=bold, size=size, color=(color or "000000"))

def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def _align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def _border_all(style="thin"):
    s = Side(style=style)
    return Border(left=s, right=s, top=s, bottom=s)

def _border_bottom(style="thin"):
    s = Side(style=style)
    return Border(bottom=s)

def _thb(val):
    if val is None: return "— ฿"
    return f"{val:,.0f} ฿"

def _pct(val):
    if val is None: return "—%"
    return f"{int(round(val * 100))}%"

def _num_or_none(val):
    """Best-effort numeric parser (788.37 / 788,37 / 1,200.50 / 1.200,50 / 1 200,50)."""
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return float(val)
    if isinstance(val, str):
        v = (
            val.strip()
            .replace(" ", "")
            .replace(" ", "")
            .replace(" ", "")
        )
        v = re.sub(r"[^\d,.\-]", "", v)
        if not v:
            return None
        if "," in v and "." in v:
            if v.rfind(",") > v.rfind("."):
                v = v.replace(".", "").replace(",", ".")
            else:
                v = v.replace(",", "")
        elif "," in v:
            parts = v.split(",")
            if len(parts) == 2:
                left, right = parts
                v = f"{left}.{right}" if 0 < len(right) <= 2 else left + right
            else:
                v = "".join(parts[:-1]) + "." + parts[-1] if len(parts[-1]) <= 2 else "".join(parts)
        try:
            return float(v)
        except ValueError:
            return None
    return None

def _num(val, default=0.0):
    parsed = _num_or_none(val)
    return default if parsed is None else parsed

def _row_value(row, idx, default=None):
    return row[idx] if len(row) > idx else default

def _ratio_or_none(val):
    """Accept both 0.087 and 8.7 -> 0.087."""
    num = _num_or_none(val)
    if num is None:
        return None
    if abs(num) > 1:
        return num / 100
    return num

def _round_money(val):
    return round(_num(val), 2)

def _property_purchase_price(info):
    for key in (
        "asset_purchase_price",
        "property_purchase_price",
        "unit_purchase_price",
        "purchase_price",
        "property_value",
        "asset_value",
        "estimated_property_value",
        "investment_value",
    ):
        value = _num_or_none(info.get(key))
        if value is not None:
            return value
    return None

def _tl_share(info):
    """TL share of NET after channel. Accept 25 or 0.25; default 0.25."""
    for key in ("tl_share_pct", "commission_rate"):
        r = _ratio_or_none(info.get(key))
        if r is not None and r > 0:
            return r
    return 0.25

def _kpi_block(ws, row, col, label, value, subtitle, bg=NAVY):
    ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col+1)
    ws.merge_cells(start_row=row+1, start_column=col, end_row=row+1, end_column=col+1)
    ws.merge_cells(start_row=row+2, start_column=col, end_row=row+2, end_column=col+1)
    lc = ws.cell(row=row, column=col, value=label)
    lc.font = _f(bold=True, size=9, color=WHITE); lc.fill = _fill(bg); lc.alignment = _align("center")
    vc = ws.cell(row=row+1, column=col, value=value)
    vc.font = _f(bold=True, size=14, color=WHITE); vc.fill = _fill(bg); vc.alignment = _align("center")
    sc = ws.cell(row=row+2, column=col, value=subtitle)
    sc.font = _f(bold=False, size=8, color=WHITE); sc.fill = _fill(bg); sc.alignment = _align("center")

def _sheet_list(sheetnames):
    return ", ".join(sheetnames) if sheetnames else "(none)"

def _expenses_sheet_name(wb):
    for name in EXPENSES_SHEET_ALIASES:
        if name in wb.sheetnames:
            return name
    return None

def _validate_input_workbook_shape(wb, workbook_label="workbook"):
    missing = [
        name for name in REQUIRED_INPUT_SHEETS
        if name not in wb.sheetnames and not (name == "Owner_Expenses" and _expenses_sheet_name(wb))
    ]
    if not missing:
        return
    msg = [
        f"Invalid InputData workbook '{workbook_label}'",
        f"missing required sheet(s): {', '.join(missing)}",
        f"found sheet(s): {_sheet_list(wb.sheetnames)}",
    ]
    if set(OWNER_REPORT_SHEETS).issubset(set(wb.sheetnames)):
        msg.append("the uploaded workbook looks like a generated OwnerReport, not an InputData file")
    msg.append("expected TL_[CODE]_InputData_YYYY-MM.xlsx with 8 required sheets (new model 25/75)")
    raise ValueError(". ".join(msg) + ".")


# ── DATA READING ──────────────────────────────────────────────────────────────
def read_input(path):
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
    except Exception as exc:
        raise ValueError(f"Could not open workbook '{os.path.basename(path)}': {exc}") from exc

    _validate_input_workbook_shape(wb, os.path.basename(path))
    data = {}

    # Property_Info — vertical: column A = key, column B = value
    ws = wb["Property_Info"]
    pi = {}
    for row in ws.iter_rows(values_only=True):
        if not any(row): continue
        key = row[0]
        val = row[1] if len(row) > 1 else None
        if key and isinstance(key, str) and re.match(r'^[a-z_]+$', key.strip()):
            pi[key.strip()] = val
    if "mgmt_start_date" in pi:
        d = pi["mgmt_start_date"]
        if isinstance(d, str):
            for fmt in ('%Y-%m-%d', '%d.%m.%Y'):
                try: pi["mgmt_start_date"] = datetime.strptime(d.strip(), fmt); break
                except ValueError: pass
    data["info"] = pi

    # Reservations
    ws = wb["Reservations"]
    res = []
    headers = None
    for row in ws.iter_rows(values_only=True):
        if not any(row): continue
        if str(row[0]).startswith("NOTES"): break
        if headers is None:
            if row[0] and isinstance(row[0], str) and re.match(r'^[a-z_]+$', str(row[0]).strip()):
                headers = [str(h).strip() if h else "" for h in row]
            continue
        if len(row) < 5: continue
        r = dict(zip(headers, row))
        for df in ("checkin_date", "checkout_date"):
            r[df] = _to_date(r.get(df))
        if isinstance(r.get("checkout_date"), datetime):
            res.append(r)
    data["reservations"] = res

    # Owner_Expenses (accepts legacy sheet name "Expenses")
    exp_sheet = _expenses_sheet_name(wb)
    ws = wb[exp_sheet]
    exp = []
    headers = None
    for row in ws.iter_rows(values_only=True):
        if not any(row): continue
        if str(row[0]).startswith("NOTES"): break
        if headers is None:
            if row[0] and isinstance(row[0], str) and re.match(r'^[a-z_]+$', str(row[0]).strip()):
                headers = [str(h).strip() if h else "" for h in row]
            continue
        date_val = _to_date(row[0])
        if not date_val: continue
        r = dict(zip(headers, row))
        r["date"] = date_val
        exp.append(r)
    data["expenses"] = exp

    # Owner_Payouts (payouts + payins)
    ws = wb["Owner_Payouts"]
    pay = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or len(row) < 2: continue
        date_val = _to_date(row[0])
        if not date_val: continue
        amount = _num(_row_value(row, 1, 0))
        ptype = str(_row_value(row, 2, "") or "")
        pay.append({
            "date": date_val, "amount": amount,
            "type": ptype,
            "reference": _row_value(row, 3, "") or "",
            "description": _row_value(row, 4, "") or "",
            "is_payin": any(m in ptype.lower() for m in PAYIN_MARKERS),
        })
    data["payouts"] = pay

    # OPEX_Budget (owner-zone budget)
    ws = wb["OPEX_Budget"]
    budget = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or len(row) < 1 or not row[0]: continue
        if str(row[0]).startswith("NOTES"): break
        code = str(row[0]).strip()
        raw = row[1] if len(row) > 1 else None
        budget[code] = _num(raw)
    data["budget"] = budget

    # Cumulative (for DAP)
    ws = wb["Cumulative"]
    cumul = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or len(row) < 2 or not row[0]: continue
        if str(row[0]).startswith("NOTES"): break
        cumul[str(row[0]).strip()] = row[1]
    data["cumulative"] = cumul

    wb.close()
    return data


# ── DATE / PERIOD HELPERS ─────────────────────────────────────────────────────
def _to_date(val):
    if isinstance(val, datetime): return val
    if isinstance(val, str):
        for fmt in ('%Y-%m-%d', '%d.%m.%Y', '%Y/%m/%d'):
            try: return datetime.strptime(val.strip(), fmt)
            except ValueError: pass
    return None

def _to_bool(val, default=False):
    if val is None:
        return default
    if isinstance(val, bool):
        return val
    if isinstance(val, (int, float)):
        return val != 0
    if isinstance(val, str):
        v = val.strip().lower()
        if v in {"1", "true", "yes", "y", "on", "strict", "block"}:
            return True
        if v in {"0", "false", "no", "n", "off", "warn", "warning"}:
            return False
    return default

def _parse_period(info):
    p = info.get("period")
    if isinstance(p, datetime):
        return p.year, p.month
    if isinstance(p, str):
        parts = p.strip().split("-")
        if len(parts) == 2:
            return int(parts[0]), int(parts[1])
    raise ValueError(f"Cannot parse period: {p}")

def _period_str(year, month):
    return f"{year:04d}-{month:02d}"

def _next_period(year, month):
    if month == 12:
        return year + 1, 1
    return year, month + 1

def _month_days(year, month):
    return calendar.monthrange(year, month)[1]

def _iter_months(start_date, end_year, end_month):
    y, m = start_date.year, start_date.month
    while (y, m) <= (end_year, end_month):
        yield y, m
        m += 1
        if m > 12:
            m = 1; y += 1

def _month_label(year, month):
    return f"{MONTHS_SHORT[month]}'{str(year)[2:]}"


# ── PER-BOOKING ECONOMICS (new model) ─────────────────────────────────────────
def _is_ota_channel(channel):
    ch = str(channel or "").lower()
    return any(m in ch for m in OTA_CHANNEL_MARKERS)

def booking_econ(r, tl_share):
    """
    Resolve the 25/75-after-channel chain for one reservation row.

    channel fee priority:
    1) channel_commission — ACTUAL amount withheld by the channel;
    2) channel_commission_pct × gross — when only the percent is known;
    3) 0 with fee_missing flag when the channel looks like OTA/agent.
    """
    gross = _num(r.get("gross_amount"))
    fee = _num_or_none(r.get("channel_commission"))
    if fee is None:
        fee = _num_or_none(r.get("ota_commission"))  # legacy column name
    fee_missing = False
    if fee is None:
        pct = _ratio_or_none(r.get("channel_commission_pct"))
        if pct is not None and pct > 0:
            fee = gross * pct
        else:
            fee = 0.0
            if _is_ota_channel(r.get("channel")):
                fee_missing = True
    net = gross - fee
    tl = net * tl_share
    owner = net - tl
    return {
        "gross": gross,
        "channel_fee": fee,
        "net": net,
        "tl": tl,
        "owner": owner,
        "fee_missing": fee_missing,
    }


# ── MONTHLY AGGREGATION ───────────────────────────────────────────────────────
def compute_monthly(data, mgmt_start, rpt_year, rpt_month):
    """
    For every month from mgmt_start to rpt_year/rpt_month compute the new-model
    chain. Revenue recognition: CHECKOUT DATE.
    Balance movement: opening + owner_share + payins − owner_expenses − payouts.
    """
    info = data["info"]
    tl_share = _tl_share(info)
    opening_key = [k for k in info if "beginning" in k.lower() or "opening" in k.lower()]
    opening_bal = _num(info.get(opening_key[0], 0), 0.0) if opening_key else 0.0

    months = []
    prev_closing = opening_bal

    for yr, mo in _iter_months(mgmt_start, rpt_year, rpt_month):
        bk = [r for r in data["reservations"]
              if r["checkout_date"].year == yr and r["checkout_date"].month == mo]
        econ = [booking_econ(r, tl_share) for r in bk]
        gross        = sum(e["gross"] for e in econ)
        channel_fees = sum(e["channel_fee"] for e in econ)
        net_channel  = sum(e["net"] for e in econ)
        tl_amt       = sum(e["tl"] for e in econ)
        owner_amt    = sum(e["owner"] for e in econ)
        bookings_count  = len(bk)
        nights_occupied = sum(_num(r.get("nights")) for r in bk)

        ex = [e for e in data["expenses"]
              if e["date"].year == yr and e["date"].month == mo]
        owner_exp = sum(_num(e.get("amount")) for e in ex)

        py_rows = [p for p in data["payouts"]
                   if p["date"].year == yr and p["date"].month == mo]
        payouts = sum(p["amount"] for p in py_rows if not p["is_payin"])
        payins  = sum(p["amount"] for p in py_rows if p["is_payin"])

        net_owner_income = owner_amt - owner_exp
        closing_bal = prev_closing + owner_amt + payins - owner_exp - payouts

        days_in_month = _month_days(yr, mo)
        occupancy = (nights_occupied / days_in_month) if days_in_month > 0 else 0
        adr = (gross / nights_occupied) if nights_occupied > 0 else 0

        months.append({
            "year": yr, "month": mo,
            "label": _month_label(yr, mo),
            "gross": gross, "channel_fees": channel_fees, "net_channel": net_channel,
            "tl_share": tl_amt, "owner_share": owner_amt,
            "owner_exp": owner_exp, "payouts": payouts, "owner_payin": payins,
            "net_owner_income": net_owner_income,
            "opening_bal": prev_closing, "closing_bal": closing_bal,
            "bookings": bookings_count, "nights": nights_occupied,
            "occupancy": occupancy, "adr": adr,
        })
        prev_closing = closing_bal

    return months


# ── SHEET 1: DASHBOARD ────────────────────────────────────────────────────────
def build_dashboard(wb, data, rpt_year, rpt_month, cur_month):
    ws = wb.create_sheet("Dashboard")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 2
    widths = {"B": 16, "C": 14, "D": 16, "E": 14, "F": 16, "G": 12, "H": 13, "I": 13}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    info = data["info"]
    tl_share  = _tl_share(info)
    prop_name  = info.get("property_name", "")
    prop_code  = info.get("property_code", "")
    owner_name = info.get("owner_name", "")
    prop_type  = info.get("property_type", "")
    bedrooms   = info.get("bedrooms", "")
    location   = info.get("location", "")

    period_label = f"{MONTHS_RU[rpt_month]} {rpt_year}"
    m = cur_month
    bk = [r for r in data["reservations"]
          if r["checkout_date"].year == rpt_year and r["checkout_date"].month == rpt_month]
    econ = {id(r): booking_econ(r, tl_share) for r in bk}

    days_in_month = _month_days(rpt_year, rpt_month)
    tl_pct_lbl = f"{tl_share*100:.0f}%"
    owner_pct_lbl = f"{(1-tl_share)*100:.0f}%"

    # Title block
    ws.row_dimensions[1].height = 6
    ws.row_dimensions[2].height = 28
    ws.merge_cells("B2:I2")
    c = ws["B2"]
    c.value = "TROPICLOOK — OWNER FINANCIAL REPORT"
    c.font = _f(bold=True, size=14, color=WHITE)
    c.fill = _fill(NAVY); c.alignment = _align("center")

    ws.row_dimensions[3].height = 18
    ws.merge_cells("B3:I3")
    c = ws["B3"]
    c.value = f"{prop_name}  |  {period_label}  |  {prop_code}"
    c.font = _f(bold=True, size=11, color=WHITE)
    c.fill = _fill(NAVY); c.alignment = _align("center")

    ws.row_dimensions[4].height = 16
    ws.merge_cells("B4:I4")
    c = ws["B4"]
    c.value = (f"Owner: {owner_name}  |  Модель: TropicLook {tl_pct_lbl} / "
               f"Собственник {owner_pct_lbl} после расходов канала  |  {bedrooms}BR {prop_type}, {location}")
    c.font = _f(bold=False, size=9, color=WHITE)
    c.fill = _fill(NAVY); c.alignment = _align("center")

    ws.row_dimensions[5].height = 8

    # KPI section
    ws.row_dimensions[6].height = 16
    ws.merge_cells("B6:I6")
    c = ws["B6"]
    c.value = "KEY PERFORMANCE INDICATORS"
    c.font = _f(bold=True, size=10, color=WHITE)
    c.fill = _fill(TEAL); c.alignment = _align("center")

    # 9 KPI blocks in 3 rows: full calculation chain
    for r in range(7, 10):
        ws.row_dimensions[r].height = 20
    _kpi_block(ws, 7, 2, "GROSS REVENUE",     _thb(m["gross"]),        "Доход от проживания до канала", NAVY)
    _kpi_block(ws, 7, 4, "CHANNEL COSTS",     _thb(-m["channel_fees"]) if m["channel_fees"] else _thb(0), "Комиссии OTA / агентов (факт)", NAVY)
    _kpi_block(ws, 7, 6, "NET ПОСЛЕ КАНАЛА",  _thb(m["net_channel"]),  "База распределения 25/75", NAVY)
    ws.merge_cells("H7:I9")  # spacer

    for r in range(10, 13):
        ws.row_dimensions[r].height = 20
    _kpi_block(ws, 10, 2, f"TROPICLOOK {tl_pct_lbl}",   _thb(-m["tl_share"]),   "Комиссия управления от NET", TEAL)
    _kpi_block(ws, 10, 4, f"OWNER SHARE {owner_pct_lbl}", _thb(m["owner_share"]), "Доля собственника от NET", TEAL)
    _kpi_block(ws, 10, 6, "OWNER EXPENSES",   _thb(-m["owner_exp"]) if m["owner_exp"] else _thb(0), "Личное / актив / общедомовое", TEAL)

    for r in range(13, 16):
        ws.row_dimensions[r].height = 20
    _kpi_block(ws, 13, 2, "NET OWNER INCOME", _thb(m["net_owner_income"]), "Чистый доход собственника", GOLD)
    _kpi_block(ws, 13, 4, "CASH BALANCE",     _thb(m["closing_bal"]),      "Остаток на счёте", GOLD)
    _kpi_block(ws, 13, 6, "OCCUPANCY / ADR",
               f"{int(round(m['occupancy']*100))}% / {m['adr']:,.0f} ฿",
               f"{int(m['nights'])} из {days_in_month} ночей", GOLD)

    ws.row_dimensions[16].height = 8

    # Bookings table
    ws.row_dimensions[17].height = 16
    ws.merge_cells("B17:I17")
    c = ws["B17"]
    c.value = f"BOOKINGS — {MONTHS_RU[rpt_month].upper()} {rpt_year}"
    c.font = _f(bold=True, size=10, color=WHITE)
    c.fill = _fill(TEAL); c.alignment = _align("left")

    ws.row_dimensions[18].height = 14
    headers = ["Бронирование", "Канал", "Гость", "Даты", "Ночей",
               "Gross (฿)", "Канал (฿)", "NET (฿)"]
    for i, h in enumerate(headers, 2):
        c = ws.cell(row=18, column=i, value=h)
        c.font = _f(bold=True, size=9, color=WHITE)
        c.fill = _fill(NAVY); c.alignment = _align("center")

    ri = 18
    for r in bk:
        ri += 1
        e = econ[id(r)]
        ws.row_dimensions[ri].height = 14
        ci = r["checkin_date"].strftime("%d.%m") if isinstance(r["checkin_date"], datetime) else ""
        co = r["checkout_date"].strftime("%d.%m") if isinstance(r["checkout_date"], datetime) else ""
        row_vals = [
            r.get("booking_id", ""), r.get("channel", ""), r.get("guest_name", ""),
            f"{ci}—{co}", int(_num(r.get("nights"))),
            round(e["gross"], 2), round(-e["channel_fee"], 2) if e["channel_fee"] else 0,
            round(e["net"], 2),
        ]
        fill = _fill(LIGHT_BG) if ri % 2 == 0 else _fill(WHITE)
        for i, v in enumerate(row_vals, 2):
            c = ws.cell(row=ri, column=i, value=v)
            c.font = _f(size=9); c.fill = fill
            c.alignment = _align("center" if i >= 5 else "left")
            if i >= 7: c.number_format = '#,##0'
    # totals row
    ri += 1
    ws.row_dimensions[ri].height = 15
    totals = ["ИТОГО", "", "", "", int(m["nights"]),
              round(m["gross"], 2), round(-m["channel_fees"], 2) if m["channel_fees"] else 0,
              round(m["net_channel"], 2)]
    for i, v in enumerate(totals, 2):
        c = ws.cell(row=ri, column=i, value=v)
        c.font = _f(bold=True, size=9); c.fill = _fill(LIGHT_BG)
        c.border = _border_bottom("medium")
        c.alignment = _align("center" if i >= 5 else "left")
        if i >= 7: c.number_format = '#,##0'

    # Model footnote
    ri += 2
    ws.merge_cells(start_row=ri, start_column=2, end_row=ri, end_column=9)
    c = ws.cell(ri, 2,
        "Модель 25/75 после расходов канала: стандартные гостевые уборка, стирка, химия, "
        "приветственные пакеты и гостевое электричество — внутри экономики TropicLook и "
        "Собственнику не выставляются. Доход признаётся в месяц выезда гостя (checkout).")
    c.font = _f(size=8, color="666666"); c.alignment = _align("left", wrap=True)
    ws.row_dimensions[ri].height = 26


# ── SHEET 2: P&L MONTHLY ──────────────────────────────────────────────────────
def build_pl(wb, data, rpt_year, rpt_month, cur_month):
    ws = wb.create_sheet("P&L Monthly")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 46
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 34

    info      = data["info"]
    tl_share  = _tl_share(info)
    prop_name = info.get("property_name", "")
    m         = cur_month
    opening   = m["opening_bal"]
    period_label = f"{MONTHS_RU[rpt_month]} {rpt_year}"
    tl_pct_lbl = f"{tl_share*100:.0f}%"
    owner_pct_lbl = f"{(1-tl_share)*100:.0f}%"

    bk = [r for r in data["reservations"]
          if r["checkout_date"].year == rpt_year and r["checkout_date"].month == rpt_month]
    ex = [e for e in data["expenses"]
          if e["date"].year == rpt_year and e["date"].month == rpt_month]
    py = [p for p in data["payouts"]
          if p["date"].year == rpt_year and p["date"].month == rpt_month]
    payout_rows = [p for p in py if not p["is_payin"]]
    payin_rows  = [p for p in py if p["is_payin"]]

    exp_by_cat = defaultdict(float)
    for e in ex:
        cat = str(e.get("category_code") or "MISC").strip()
        exp_by_cat[cat] += _num(e.get("amount"))

    row = 1
    def next_row():
        nonlocal row; row += 1; return row

    def write(r, label, val=None, note=None, bold=False, indent=False,
              bg=None, fg="000000", size=10, num_fmt='#,##0.00'):
        lbl = ("   " if indent else "") + (label or "")
        ws.row_dimensions[r].height = 15
        c_label = ws.cell(r, 2, lbl)
        c_label.font = _f(bold=bold, size=size, color=fg)
        if bg: c_label.fill = _fill(bg)
        c_label.alignment = _align("left")
        if val is not None:
            if isinstance(val, float):
                val = round(val, 2)
            c_val = ws.cell(r, 3, val)
            c_val.font = _f(bold=bold, size=size, color=fg)
            if bg: c_val.fill = _fill(bg)
            c_val.alignment = _align("right")
            if num_fmt: c_val.number_format = num_fmt
        if note is not None:
            c_note = ws.cell(r, 4, note)
            c_note.font = _f(size=8, color="666666")
            c_note.alignment = _align("left")

    def write_section(r, label):
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
        c = ws.cell(r, 2, label)
        c.font = _f(bold=True, size=10, color=WHITE)
        c.fill = _fill(TEAL); c.alignment = _align("left")
        ws.row_dimensions[r].height = 16

    def write_total(r, label, val, note=None):
        write(r, label, val, note, bold=True, bg=LIGHT_BG)
        ws.cell(r, 2).border = _border_bottom("medium")
        ws.cell(r, 3).border = _border_bottom("medium")

    def write_navy(r, label, val, note=None, size=11):
        ws.row_dimensions[r].height = 17
        c = ws.cell(r, 2, label)
        c.font = _f(bold=True, size=size, color=WHITE)
        c.fill = _fill(NAVY); c.alignment = _align("left")
        c2 = ws.cell(r, 3, round(val, 2) if isinstance(val, float) else val)
        c2.font = _f(bold=True, size=size, color=WHITE); c2.fill = _fill(NAVY)
        c2.alignment = _align("right"); c2.number_format = '#,##0.00'
        c3 = ws.cell(r, 4, note or "")
        c3.font = _f(size=8, color=WHITE); c3.fill = _fill(NAVY); c3.alignment = _align("left")

    # Title
    ws.row_dimensions[1].height = 20
    ws.merge_cells("B1:D1")
    c = ws["B1"]; c.value = f"P&L — {prop_name} — {period_label}"
    c.font = _f(bold=True, size=12, color=WHITE); c.fill = _fill(NAVY)
    c.alignment = _align("center")

    ws.row_dimensions[2].height = 14
    ws.merge_cells("B2:D2")
    c = ws["B2"]
    c.value = (f"Цепочка: Gross → расходы канала → NET → TropicLook {tl_pct_lbl} / "
               f"Собственник {owner_pct_lbl} → расходы Собственника → выплата.  "
               f"Начальный баланс: {m['opening_bal']:,.2f} ฿")
    c.font = _f(bold=False, size=9); c.alignment = _align("left")

    ws.row_dimensions[3].height = 14
    for col, hdr in [(2, None), (3, period_label), (4, "Примечание")]:
        c = ws.cell(3, col, hdr)
        if col > 2:
            c.font = _f(bold=True, size=9, color=WHITE)
            c.fill = _fill(NAVY); c.alignment = _align("center")

    row = 3
    # A. REVENUE
    write_section(next_row(), "REVENUE / ДОХОД ОТ ПРОЖИВАНИЯ")
    for r in bk:
        e = booking_econ(r, tl_share)
        nights = int(_num(r.get("nights")))
        write(next_row(), str(r.get("booking_id", "")), e["gross"],
              f"{r.get('channel','')}, {nights}н", indent=True)
    write_total(next_row(), "ИТОГО ВАЛОВОЙ ДОХОД (A)", m["gross"], "До расходов канала")

    # B. CHANNEL COSTS
    write_section(next_row(), "CHANNEL COSTS / РАСХОДЫ КАНАЛА (OTA / АГЕНТ)")
    any_fee = False
    for r in bk:
        e = booking_econ(r, tl_share)
        if e["channel_fee"] > 0:
            any_fee = True
            pct = e["channel_fee"]/e["gross"]*100 if e["gross"] > 0 else 0
            write(next_row(), f"{r.get('booking_id','')} — комиссия канала", -e["channel_fee"],
                  f"{r.get('channel','')}, {pct:.1f}% (факт удержания)", indent=True)
    if not any_fee:
        write(next_row(), "Комиссии канала отсутствуют", 0.0,
              "Прямые продажи / возвратные гости", indent=True)
    write_total(next_row(), "ИТОГО РАСХОДЫ КАНАЛА (B)", -m["channel_fees"])

    # C. NET
    write_navy(next_row(), "NET ПОСЛЕ КАНАЛА (C = A − B)", m["net_channel"],
               f"База распределения {tl_pct_lbl}/{owner_pct_lbl}")

    # D/E. DISTRIBUTION
    write_section(next_row(), "DISTRIBUTION / РАСПРЕДЕЛЕНИЕ")
    write(next_row(), f"Комиссия TropicLook ({tl_pct_lbl} от C) (D)", -m["tl_share"],
          "Управление, гостевые операции: уборка, стирка, электричество, welcome pack", indent=False)
    write(next_row(), f"Доля Собственника ({owner_pct_lbl} от C) (E)", m["owner_share"],
          "До расходов Собственника", indent=False)

    # F. OWNER EXPENSES
    write_section(next_row(), "OWNER EXPENSES / РАСХОДЫ СОБСТВЕННИКА")
    budget = data["budget"]
    shown = set()
    for cat in BUDGET_ORDER:
        if cat not in exp_by_cat or cat in shown: continue
        shown.add(cat)
        fact = exp_by_cat[cat]
        bgt  = budget.get(cat, 0)
        note = f"бюджет: {bgt:,.0f}" if bgt else ""
        write(next_row(), CAT_NAMES.get(cat, cat), -fact, note, indent=True)
    for cat, fact in exp_by_cat.items():
        if cat not in shown:
            write(next_row(), CAT_NAMES.get(cat, cat), -fact, indent=True)
    if not exp_by_cat:
        write(next_row(), "Расходов Собственника в отчётном месяце нет", 0.0,
              "Только личное проживание / актив / общедомовое", indent=True)
    write_total(next_row(), "ИТОГО РАСХОДЫ СОБСТВЕННИКА (F)", -m["owner_exp"],
                f"{len(ex)} операций")

    # NET OWNER INCOME
    write_navy(next_row(), "ЧИСТЫЙ ДОХОД СОБСТВЕННИКА (E − F)", m["net_owner_income"],
               "К зачислению на счёт Собственника")

    # PAYIN (if any)
    if payin_rows:
        write_section(next_row(), "OWNER PAYIN / ПОСТУПЛЕНИЯ ОТ СОБСТВЕННИКА")
        for p in payin_rows:
            d = p["date"].strftime("%d.%m.%Y")
            write(next_row(), (p["description"] or p["type"])[:55], p["amount"], d, indent=True)
        write_total(next_row(), "ИТОГО ПОСТУПЛЕНИЯ (P)", m["owner_payin"])

    # G. PAYOUTS
    write_section(next_row(), "PAYOUTS / ВЫПЛАТЫ СОБСТВЕННИКУ")
    for p in payout_rows:
        d = p["date"].strftime("%d.%m.%Y")
        write(next_row(), (p["description"] or p["type"])[:55], -p["amount"], d, indent=True)
    write_total(next_row(), "ИТОГО ВЫПЛАТЫ (G)", -m["payouts"], f"{len(payout_rows)} выплат")

    # BALANCE MOVEMENT
    write_section(next_row(), "ДВИЖЕНИЕ БАЛАНСА")
    write(next_row(), "Баланс на начало", opening)
    write(next_row(), f"+ Доля Собственника {owner_pct_lbl} (E)", m["owner_share"])
    if m["owner_payin"]:
        write(next_row(), "+ Поступления от Собственника (P)", m["owner_payin"])
    write(next_row(), "− Расходы Собственника (F)", -m["owner_exp"])
    write(next_row(), "− Выплаты (G)", -m["payouts"])

    final_row = next_row()
    ws.row_dimensions[final_row].height = 18
    c = ws.cell(final_row, 2, "БАЛАНС НА КОНЕЦ")
    c.font = _f(bold=True, size=12, color=WHITE); c.fill = _fill(NAVY)
    c.alignment = _align("left")
    c2 = ws.cell(final_row, 3, round(m["closing_bal"], 2))
    c2.font = _f(bold=True, size=12, color=WHITE); c2.fill = _fill(NAVY)
    c2.alignment = _align("right"); c2.number_format = '#,##0.00'
    ws.cell(final_row, 4).fill = _fill(NAVY)

    # Reference footnote: what is inside TL economics
    fr = next_row() + 1
    ws.merge_cells(start_row=fr, start_column=2, end_row=fr, end_column=4)
    c = ws.cell(fr, 2,
        "Справочно: стандартные гостевые уборка, стирка, химия, приветственные пакеты и "
        "гостевое электричество включены в экономику TropicLook (комиссия D) и в расходы "
        "Собственника не входят. Resort Fee — отдельный сбор TropicLook, в расчёт с "
        "Собственником не включается.")
    c.font = _f(size=8, color="666666"); c.alignment = _align("left", wrap=True)
    ws.row_dimensions[fr].height = 30


# ── SHEET 3: OPEX PASSPORT (owner-zone only) ─────────────────────────────────
def build_opex_passport(wb, data, rpt_year, rpt_month):
    ws = wb.create_sheet("OPEX Passport")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 36
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 10
    ws.column_dimensions["G"].width = 16

    ws.row_dimensions[1].height = 20
    ws.merge_cells("B1:G1")
    c = ws["B1"]; c.value = f"OWNER EXPENSES PASSPORT — {MONTHS_RU[rpt_month]} {rpt_year}"
    c.font = _f(bold=True, size=12, color=WHITE); c.fill = _fill(NAVY)
    c.alignment = _align("center")

    ws.row_dimensions[2].height = 14
    ws.merge_cells("B2:G2")
    c = ws["B2"]
    c.value = ("Только расходы зоны Собственника (личное проживание / актив / общедомовое). "
               "Стандартные гостевые операционные расходы — в экономике TropicLook.")
    c.font = _f(size=8, color="666666"); c.alignment = _align("left")

    ws.row_dimensions[3].height = 14
    hdrs = ["Категория", "Бюджет", "Факт", "Δ", "Δ%", "Статус"]
    for i, h in enumerate(hdrs, 2):
        c = ws.cell(3, i, h)
        c.font = _f(bold=True, size=9, color=WHITE)
        c.fill = _fill(TEAL); c.alignment = _align("center")

    budget = data["budget"]
    ex = [e for e in data["expenses"]
          if e["date"].year == rpt_year and e["date"].month == rpt_month]
    exp_by_cat = defaultdict(float)
    for e in ex:
        cat = str(e.get("category_code") or "MISC").strip()
        exp_by_cat[cat] += _num(e.get("amount"))

    all_cats = []
    for cat in BUDGET_ORDER:
        if cat not in all_cats:
            all_cats.append(cat)
    for cat in list(exp_by_cat) + list(budget):
        if cat not in all_cats:
            all_cats.append(cat)

    total_bgt = total_fact = 0
    row = 3
    seen_names = set()
    for cat in all_cats:
        fact = exp_by_cat.get(cat, 0)
        bgt  = budget.get(cat, 0) or 0
        if fact == 0 and bgt == 0:
            continue
        name = CAT_NAMES.get(cat, cat)
        if name in seen_names and fact == 0:
            continue
        seen_names.add(name)
        row += 1
        ws.row_dimensions[row].height = 14
        delta = fact - bgt
        delta_pct = (delta / bgt) if bgt != 0 else (-1 if fact == 0 else 0)
        if bgt == 0 and fact > 0:
            status = "⚠️ Внеплановое"
        elif bgt > 0 and fact > bgt * 1.30:
            status = "🔴 Перерасход"
        elif bgt > 0 and fact > bgt * 1.10:
            status = "⚠️ Превышение"
        else:
            status = "✅"
        fill = _fill(LIGHT_BG) if row % 2 == 0 else _fill(WHITE)
        vals = [name, bgt, fact, delta, delta_pct, status]
        for i, v in enumerate(vals, 2):
            c = ws.cell(row, i, v)
            c.font = _f(size=9); c.fill = fill; c.alignment = _align("center" if i >= 3 else "left")
            if i in (3, 4, 5): c.number_format = '#,##0'
            if i == 6: c.number_format = '0.0%'
        total_bgt += bgt; total_fact += fact

    row += 1
    ws.row_dimensions[row].height = 15
    for i, v in enumerate(["ИТОГО", total_bgt, total_fact, total_fact - total_bgt, None, None], 2):
        c = ws.cell(row, i, v)
        c.font = _f(bold=True, size=10); c.fill = _fill(LIGHT_BG)
        c.border = _border_bottom("medium"); c.alignment = _align("center" if i >= 3 else "left")
        if i in (3, 4, 5): c.number_format = '#,##0'


# ── SHEET 4: 12-MONTH SUMMARY ─────────────────────────────────────────────────
def build_12month(wb, data, rpt_year, rpt_month, months):
    ws = wb.create_sheet("12-Month Summary")
    ws.sheet_view.showGridLines = False

    info      = data["info"]
    tl_share  = _tl_share(info)
    prop_name = info.get("property_name", "")
    mgmt_start = info.get("mgmt_start_date")
    if isinstance(mgmt_start, datetime):
        start_str = f"{MONTHS_RU[mgmt_start.month]} {mgmt_start.year}"
    else:
        start_str = "—"
    end_str = f"{MONTHS_RU[rpt_month]} {rpt_year}"
    tl_pct_lbl = f"{tl_share*100:.0f}%"
    owner_pct_lbl = f"{(1-tl_share)*100:.0f}%"

    num_months = len(months)
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 26
    for i in range(num_months):
        col = get_column_letter(3 + i)
        ws.column_dimensions[col].width = 13
    total_col = get_column_letter(3 + num_months)
    ws.column_dimensions[total_col].width = 14

    def mc(r, c): return ws.cell(r, c)

    ws.row_dimensions[1].height = 20
    last_col = 3 + num_months
    ws.merge_cells(start_row=1, start_column=2, end_row=1, end_column=last_col)
    c = ws["B1"]; c.value = f"PERFORMANCE OVERVIEW — {prop_name}"
    c.font = _f(bold=True, size=12, color=WHITE); c.fill = _fill(NAVY); c.alignment = _align("center")

    ws.row_dimensions[2].height = 14
    ws.merge_cells(start_row=2, start_column=2, end_row=2, end_column=last_col)
    c = ws["B2"]; c.value = f"{start_str} — {end_str} (management period), модель {tl_pct_lbl}/{owner_pct_lbl} после расходов канала"
    c.font = _f(size=9); c.alignment = _align("center")

    ws.row_dimensions[3].height = 14
    mc(3, 2).value = "Показатель"
    mc(3, 2).font = _f(bold=True, size=9, color=WHITE); mc(3, 2).fill = _fill(NAVY); mc(3, 2).alignment = _align("left")
    for i, mo in enumerate(months):
        c = mc(3, 3 + i); c.value = mo["label"]
        c.font = _f(bold=True, size=9, color=WHITE); c.fill = _fill(NAVY); c.alignment = _align("center")
    tc = mc(3, last_col); tc.value = "TOTAL"
    tc.font = _f(bold=True, size=9, color=WHITE); tc.fill = _fill(NAVY); tc.alignment = _align("center")

    def write_row(r, label, values, totals_fn=sum, bold=False, bg=None, is_section=False):
        ws.row_dimensions[r].height = 14
        if is_section:
            ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=last_col)
            c = mc(r, 2); c.value = label
            c.font = _f(bold=True, size=9, color=WHITE); c.fill = _fill(TEAL); c.alignment = _align("left")
            return
        fill = _fill(bg or WHITE)
        c = mc(r, 2); c.value = label
        c.font = _f(bold=bold, size=9); c.fill = fill; c.alignment = _align("left")
        non_none = [v for v in values if v is not None]
        total = totals_fn(non_none) if non_none else None
        for i, v in enumerate(values):
            c = mc(r, 3 + i)
            c.value = round(v, 2) if isinstance(v, float) else v
            c.font = _f(bold=bold, size=9); c.fill = fill; c.alignment = _align("right")
        tc = mc(r, last_col)
        tc.value = round(total, 2) if isinstance(total, float) else total
        tc.font = _f(bold=bold, size=9); tc.fill = _fill(LIGHT_BG) if not bg else fill
        tc.alignment = _align("right")

    row = 3
    row += 1; write_row(row, "REVENUE", [], is_section=True)
    row += 1; write_row(row, "Gross Revenue (฿)",      [m["gross"] for m in months], bg=LIGHT_BG)
    row += 1; write_row(row, "Channel costs OTA (฿)",  [m["channel_fees"] for m in months])
    row += 1; write_row(row, "NET после канала (฿)",   [m["net_channel"] for m in months], bold=True, bg=LIGHT_BG)
    row += 1; write_row(row, "Bookings (#)",           [m["bookings"] for m in months])
    row += 1; write_row(row, "Nights",                 [int(m["nights"]) for m in months], bg=LIGHT_BG)
    row += 1; write_row(row, "DISTRIBUTION", [], is_section=True)
    row += 1; write_row(row, f"TropicLook {tl_pct_lbl} (฿)",   [m["tl_share"] for m in months])
    row += 1; write_row(row, f"Owner Share {owner_pct_lbl} (฿)", [m["owner_share"] for m in months], bold=True, bg=LIGHT_BG)
    row += 1; write_row(row, "OWNER EXPENSES", [], is_section=True)
    row += 1; write_row(row, "Owner Expenses (฿)",     [m["owner_exp"] for m in months])
    row += 1; write_row(row, "RESULT", [], is_section=True)
    row += 1; write_row(row, "Net Owner Income (฿)",   [m["net_owner_income"] for m in months], bold=True)
    row += 1; write_row(row, "Owner Payin (฿)",        [m["owner_payin"] for m in months], bg=LIGHT_BG)
    row += 1; write_row(row, "Payouts (฿)",            [m["payouts"] for m in months])
    row += 1; write_row(row, "Closing Balance (฿)",    [m["closing_bal"] for m in months],
                        totals_fn=lambda x: x[-1] if x else 0, bold=True, bg=LIGHT_BG)
    row += 1; write_row(row, "KPIs", [], is_section=True)
    row += 1; write_row(row, "Occupancy %",
                        [int(round(m["occupancy"]*100)) for m in months],
                        totals_fn=lambda x: round(sum(x)/len(x)) if x else 0)
    row += 1; write_row(row, "ADR (฿)",
                        [int(round(m["adr"])) for m in months],
                        totals_fn=lambda x: round(sum(x)/len(x)) if x else 0, bg=LIGHT_BG)


# ── SHEET 5: TRANSACTION LEDGER ───────────────────────────────────────────────
def build_ledger(wb, data, rpt_year, rpt_month, cur_month):
    ws = wb.create_sheet("Transaction Ledger")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 7
    ws.column_dimensions["C"].width = 48
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 14
    ws.column_dimensions["G"].width = 14

    info      = data["info"]
    tl_share  = _tl_share(info)
    prop_name = info.get("property_name", "")
    opening   = cur_month["opening_bal"]
    period_label = f"{MONTHS_RU[rpt_month]} {rpt_year}"
    tl_pct_lbl = f"{tl_share*100:.0f}%"

    ws.row_dimensions[1].height = 20
    ws.merge_cells("B1:G1")
    c = ws["B1"]; c.value = f"OWNER ACCOUNT LEDGER — {prop_name} — {period_label}"
    c.font = _f(bold=True, size=12, color=WHITE); c.fill = _fill(NAVY); c.alignment = _align("center")

    ws.row_dimensions[2].height = 14
    for i, h in enumerate(["Дата", "Описание", "Кат.", "Приход", "Расход", "Баланс"], 2):
        c = ws.cell(2, i, h)
        c.font = _f(bold=True, size=9, color=WHITE)
        c.fill = _fill(TEAL); c.alignment = _align("center")

    transactions = []

    # Owner expenses (dated)
    ex = [e for e in data["expenses"]
          if e["date"].year == rpt_year and e["date"].month == rpt_month]
    for e in ex:
        cat = str(e.get("category_code") or "MISC")
        transactions.append({
            "date": e["date"],
            "desc": (e.get("description") or "")[:55],
            "cat":  cat,
            "debit": _num(e.get("amount")),
            "credit": 0,
            "sort": 4,
        })

    # Bookings: gross + channel fee + TL share on checkout date
    bk = [r for r in data["reservations"]
          if r["checkout_date"].year == rpt_year and r["checkout_date"].month == rpt_month]
    for r in bk:
        co  = r["checkout_date"]
        bid = str(r.get("booking_id", ""))
        e   = booking_econ(r, tl_share)
        if e["gross"] > 0:
            transactions.append({"date": co, "desc": f"Gross: {bid} (выезд — checkout)",
                                  "cat": "REVENUE", "credit": round(e["gross"], 2), "debit": 0, "sort": 1})
        if e["channel_fee"] > 0:
            transactions.append({"date": co, "desc": f"Комиссия канала: {bid}",
                                  "cat": "CHANNEL-FEE", "debit": round(e["channel_fee"], 2), "credit": 0, "sort": 2})
        if e["tl"] > 0:
            transactions.append({"date": co, "desc": f"TropicLook {tl_pct_lbl} от NET: {bid}",
                                  "cat": "MGMT-FEE", "debit": round(e["tl"], 2), "credit": 0, "sort": 3})

    # Payouts / payins
    py = [p for p in data["payouts"]
          if p["date"].year == rpt_year and p["date"].month == rpt_month]
    for p in py:
        if p["is_payin"]:
            transactions.append({
                "date": p["date"],
                "desc": (p["description"] or "Поступление от Собственника")[:55],
                "cat": "PAYIN", "credit": p["amount"], "debit": 0, "sort": 5,
            })
        else:
            transactions.append({
                "date": p["date"],
                "desc": (p["description"] or "Выплата Собственнику")[:55],
                "cat": "PAYOUT", "debit": p["amount"], "credit": 0, "sort": 6,
            })

    transactions.sort(key=lambda x: (x["date"], x["sort"]))

    row = 2

    def ledger_row(r, date_str, desc, cat, credit, debit, balance, bold=False, bg=None):
        ws.row_dimensions[r].height = 14
        fill = _fill(bg) if bg else (_fill(LIGHT_BG) if r % 2 == 0 else _fill(WHITE))
        vals = [date_str, desc, cat, credit or None, debit or None, balance]
        for i, v in enumerate(vals, 2):
            c = ws.cell(r, i, v)
            c.font = _f(bold=bold, size=9)
            c.fill = fill
            c.alignment = _align("right" if i >= 5 else ("center" if i == 2 else "left"))
            if i >= 5 and v is not None:
                c.number_format = '#,##0.00'

    row += 1
    ws.row_dimensions[row].height = 15
    ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=4)
    c = ws.cell(row, 3, "НАЧАЛЬНЫЙ БАЛАНС")
    c.font = _f(bold=True, size=9, color=WHITE); c.fill = _fill(TEAL); c.alignment = _align("center")
    c2 = ws.cell(row, 7, round(opening, 2))
    c2.font = _f(bold=True, size=9, color=WHITE); c2.fill = _fill(TEAL); c2.alignment = _align("right")
    c2.number_format = '#,##0.00'

    balance = opening
    for t in transactions:
        row += 1
        balance = balance + t["credit"] - t["debit"]
        date_str = t["date"].strftime("%d.%m") if isinstance(t["date"], datetime) else ""
        ledger_row(row, date_str, t["desc"], t["cat"],
                   t["credit"] if t["credit"] > 0 else None,
                   t["debit"]  if t["debit"]  > 0 else None,
                   round(balance, 2))

    row += 1
    ws.row_dimensions[row].height = 15
    ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=4)
    c = ws.cell(row, 3, "КОНЕЧНЫЙ БАЛАНС")
    c.font = _f(bold=True, size=9, color=WHITE); c.fill = _fill(NAVY); c.alignment = _align("center")
    c2 = ws.cell(row, 7, round(balance, 2))
    c2.font = _f(bold=True, size=9, color=WHITE); c2.fill = _fill(NAVY); c2.alignment = _align("right")
    c2.number_format = '#,##0.00'

    # Footnote
    row += 2
    bk_str = "; ".join(
        f"{r.get('booking_id')}: заезд {r['checkin_date'].strftime('%d.%m')} → выезд {r['checkout_date'].strftime('%d.%m')}"
        for r in bk if isinstance(r.get("checkin_date"), datetime)
    )
    note = ("⚠ Доход признаётся в месяц выезда гостя (checkout). Комиссия канала — фактическое "
            "удержание OTA / агента. TropicLook-доля считается от NET после канала.")
    if bk_str:
        note += f" {bk_str}."
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=7)
    c = ws.cell(row, 2, note)
    c.font = _f(size=8, color="666666"); c.alignment = _align("left", wrap=True)
    ws.row_dimensions[row].height = 28


# ── SHEET 6: DAP SNAPSHOT ─────────────────────────────────────────────────────
def build_dap(wb, data, rpt_year, rpt_month, months):
    ws = wb.create_sheet("DAP Snapshot")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 24
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 4
    ws.column_dimensions["E"].width = 24
    ws.column_dimensions["F"].width = 22

    info = data["info"]
    tl_share   = _tl_share(info)
    prop_name  = info.get("property_name", "")
    prop_code  = info.get("property_code", "")
    location   = info.get("location", "")
    prop_type  = info.get("property_type", "")
    bedrooms   = info.get("bedrooms", "")
    owner_name = info.get("owner_name", "")
    mgmt_start = info.get("mgmt_start_date")
    mgmt_str   = mgmt_start.strftime("%B %Y") if isinstance(mgmt_start, datetime) else "—"
    tl_pct_lbl = f"{tl_share*100:.0f}"
    owner_pct_lbl = f"{(1-tl_share)*100:.0f}"

    period_label = f"{MONTHS_RU[rpt_month]} {rpt_year}"

    total_gross   = sum(m["gross"]        for m in months)
    total_channel = sum(m["channel_fees"] for m in months)
    total_net     = sum(m["net_channel"]  for m in months)
    total_tl      = sum(m["tl_share"]     for m in months)
    total_owner   = sum(m["owner_share"]  for m in months)
    total_exp     = sum(m["owner_exp"]    for m in months)
    total_noi     = sum(m["net_owner_income"] for m in months)
    total_pays    = sum(m["payouts"]      for m in months)
    total_bk      = sum(m["bookings"]     for m in months)
    total_nights  = sum(m["nights"]       for m in months)
    current_bal   = months[-1]["closing_bal"] if months else 0
    avg_occ       = sum(m["occupancy"] for m in months) / len(months) if months else 0
    avg_stay      = total_nights / total_bk if total_bk > 0 else 0
    exp_months    = [m["owner_exp"] for m in months]
    avg_exp       = sum(exp_months) / len(exp_months) if exp_months else 0
    channel_ratio = total_channel / total_gross if total_gross > 0 else 0

    hi_m = max(months, key=lambda m: m["owner_exp"])
    lo_m = min(months, key=lambda m: m["owner_exp"])

    purchase = _property_purchase_price(info)
    if purchase and purchase > 0 and months:
        ann_yield = (total_noi / purchase) * (12 / len(months))
        yield_str = f"{ann_yield*100:.2f}%"
    else:
        yield_str = "— (нет asset_purchase_price)"

    ws.row_dimensions[1].height = 22
    ws.merge_cells("B1:F1")
    c = ws["B1"]; c.value = f"DIGITAL ASSET PASSPORT — {prop_name}"
    c.font = _f(bold=True, size=12, color=WHITE); c.fill = _fill(NAVY); c.alignment = _align("center")

    ws.row_dimensions[2].height = 14
    ws.merge_cells("B2:F2")
    c = ws["B2"]; c.value = f"{prop_code} | ACTIVE | {period_label} | Model {tl_pct_lbl}/{owner_pct_lbl} after OTA"
    c.font = _f(size=9, color=WHITE); c.fill = _fill(NAVY); c.alignment = _align("center")

    def section_hdr(row, col, label, bg=TEAL):
        ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col+1)
        c = ws.cell(row, col, label)
        c.font = _f(bold=True, size=9, color=WHITE); c.fill = _fill(bg); c.alignment = _align("left")
        ws.row_dimensions[row].height = 14

    def kv_row(row, col, key, val):
        ws.row_dimensions[row].height = 13
        ck = ws.cell(row, col, key); ck.font = _f(size=9); ck.alignment = _align("left")
        cv = ws.cell(row, col+1, val); cv.font = _f(bold=True, size=9); cv.alignment = _align("left")

    ws.row_dimensions[3].height = 8

    section_hdr(4, 2, "L1 — IDENTITY")
    section_hdr(4, 5, "L3 — OPERATIONAL")
    r = 4
    id_data = [
        ("Property", prop_name), ("Location", location),
        ("Type", f"{prop_type} {bedrooms}BR"),
        ("Owner", owner_name),
        ("Mgmt Start", mgmt_str),
        ("Contract", f"Management {tl_pct_lbl}/{owner_pct_lbl} after OTA"),
        ("Status", "ACTIVE"),
    ]
    op_data = [
        ("Bookings (cumul)", str(int(total_bk))),
        ("Nights (cumul)", str(int(total_nights))),
        ("Avg Occupancy", f"{int(round(avg_occ*100))}%"),
        ("Avg Stay", f"{avg_stay:.1f} nights"),
        ("Channel cost ratio", f"{channel_ratio*100:.1f}% of Gross"),
        ("Annualized Yield", yield_str),
    ]
    for i, (k, v) in enumerate(id_data):
        kv_row(r + 1 + i, 2, k, v)
    for i, (k, v) in enumerate(op_data):
        kv_row(r + 1 + i, 5, k, v)

    last_id_row = r + 1 + len(id_data) + 1
    ws.row_dimensions[last_id_row].height = 6

    section_hdr(last_id_row + 1, 2, "L5 — FINANCIAL (cumul)")
    section_hdr(last_id_row + 1, 5, "L4 — OWNER EXPENSES")
    r2 = last_id_row + 1
    fin_data = [
        ("Cumul Gross Revenue",    f"{total_gross:,.0f} ฿"),
        ("Cumul Channel costs",    f"{total_channel:,.0f} ฿"),
        ("Cumul NET after OTA",    f"{total_net:,.0f} ฿"),
        (f"Cumul TropicLook {tl_pct_lbl}%", f"{total_tl:,.0f} ฿"),
        (f"Cumul Owner Share {owner_pct_lbl}%", f"{total_owner:,.0f} ฿"),
        ("Cumul Owner Expenses",   f"{total_exp:,.0f} ฿"),
        ("Cumul Net Owner Income", f"{total_noi:,.0f} ฿"),
        ("Total Payouts",          f"{total_pays:,.0f} ฿"),
        ("Current Balance",        f"{current_bal:,.0f} ฿"),
    ]
    exp_data = [
        ("Avg Monthly Owner Exp", f"{avg_exp:,.0f} ฿"),
        ("Highest month",         f"{hi_m['label']} ({hi_m['owner_exp']:,.0f} ฿)"),
        ("Lowest month",          f"{lo_m['label']} ({lo_m['owner_exp']:,.0f} ฿)"),
        ("Scope",                 "Personal stay / asset / building"),
        ("Guest opex",            "Inside TropicLook economics"),
        ("Resort Fee",            "TL stream, not in owner settlement"),
    ]
    for i, (k, v) in enumerate(fin_data):
        kv_row(r2 + 1 + i, 2, k, v)
    for i, (k, v) in enumerate(exp_data):
        kv_row(r2 + 1 + i, 5, k, v)


# ── VALIDATION (new-model rules) ─────────────────────────────────────────────
def validate(data, rpt_year, rpt_month, cur_month):
    """
    R1: closing balance >= 0 (non-blocking by default, CFO escalation).
    R2: arithmetic: opening + owner_share + payins − owner_exp − payouts = closing (±1 THB).
    R3: every owner expense has a category code (blocking).
    R3b: guest-zone codes are forbidden in owner expenses (blocking).
    R4: zero bookings — warning.
    R5: OTA-like channel without channel commission — warning per booking
        ('Direct from booking.com' is still Booking.com).
    R6: channel fee sanity: fee > gross (blocking); fee share > 30% (warning).
    """
    errors = []; warnings = []
    info = data["info"]
    tl_share = _tl_share(info)
    block_on_negative_balance = _to_bool(info.get("block_on_negative_balance"), default=False)

    m = cur_month
    bk = [r for r in data["reservations"]
          if r["checkout_date"].year == rpt_year and r["checkout_date"].month == rpt_month]
    ex = [e for e in data["expenses"]
          if e["date"].year == rpt_year and e["date"].month == rpt_month]

    # R1
    if m["closing_bal"] < -0.01:
        msg = f"R1 FAIL: Closing balance {m['closing_bal']:.2f} < 0. Escalate to CFO."
        if block_on_negative_balance:
            errors.append(msg)
        else:
            warnings.append(msg.replace("R1 FAIL", "R1 WARN") + " Report generated in non-blocking mode.")

    # R2
    calc_closing = (m["opening_bal"] + m["owner_share"] + m["owner_payin"]
                    - m["owner_exp"] - m["payouts"])
    if abs(calc_closing - m["closing_bal"]) > 1.0:
        errors.append(f"R2 FAIL: Balance discrepancy {abs(calc_closing - m['closing_bal']):.2f} THB > 1 THB.")

    # R3 / R3b
    no_cat = [e for e in ex if not e.get("category_code")]
    if no_cat:
        errors.append(f"R3 FAIL: {len(no_cat)} owner expense(s) missing category code.")
    forbidden_rows = []
    for e in ex:
        code = str(e.get("category_code") or "").strip()
        if code in FORBIDDEN_GUEST_CODES:
            forbidden_rows.append((e["date"].strftime("%Y-%m-%d"), code,
                                   FORBIDDEN_GUEST_CODES[code]))
    if forbidden_rows:
        detail = "; ".join(f"{d} {c} ({why})" for d, c, why in forbidden_rows[:5])
        errors.append(
            f"R3b FAIL: {len(forbidden_rows)} expense(s) use guest-zone codes forbidden in the "
            f"new model 25/75 after OTA: {detail}. Standard guest cleaning / laundry / chemicals / "
            f"welcome packs / guest electricity are inside TropicLook economics and must not be "
            f"charged to the Owner."
        )

    # R4
    if len(bk) == 0:
        warnings.append("R4 WARN: 0 bookings in report period — verify against PMS.")

    # R5 / R6
    for r in bk:
        e = booking_econ(r, tl_share)
        bid = r.get("booking_id", "")
        if e["fee_missing"]:
            warnings.append(
                f"R5 WARN: booking {bid} ({r.get('channel','')}) has no channel commission. "
                f"'Direct from booking.com' is still Booking.com — verify the actual amount withheld."
            )
        if e["gross"] > 0 and e["channel_fee"] > e["gross"]:
            errors.append(f"R6 FAIL: booking {bid} channel fee {e['channel_fee']:.2f} exceeds gross {e['gross']:.2f}.")
        elif e["gross"] > 0 and e["channel_fee"] / e["gross"] > 0.30:
            warnings.append(
                f"R6 WARN: booking {bid} channel fee is {e['channel_fee']/e['gross']*100:.1f}% of gross (>30%) — verify."
            )

    return errors, warnings


# ── NEXT INPUT TEMPLATE ROLLFORWARD ──────────────────────────────────────────
def _find_key_row(ws, key, key_col=1):
    for row in range(1, ws.max_row + 1):
        if ws.cell(row=row, column=key_col).value == key:
            return row
    raise KeyError(f"Key '{key}' not found in sheet '{ws.title}'.")

def _last_metric_row(ws, key_col=1):
    last = 0
    for row in range(1, ws.max_row + 1):
        val = ws.cell(row=row, column=key_col).value
        if isinstance(val, str) and re.match(r"^[a-z_]+$", val.strip()):
            last = row
    return last or ws.max_row

def _ensure_kv(ws, key, value=None, note=None, key_col=1, value_col=2):
    try:
        row = _find_key_row(ws, key, key_col=key_col)
    except KeyError:
        row = _last_metric_row(ws, key_col=key_col) + 1
        ws.insert_rows(row)
        ws.cell(row=row, column=key_col).value = key
    if value is not None:
        ws.cell(row=row, column=value_col).value = value
    if note is not None:
        ws.cell(row=row, column=value_col + 1).value = note
    return row

def _set_kv(ws, key, value, key_col=1, value_col=2, num_fmt=None, create=True):
    try:
        row = _find_key_row(ws, key, key_col=key_col)
    except KeyError:
        if not create:
            return
        row = _ensure_kv(ws, key)
    cell = ws.cell(row=row, column=value_col)
    cell.value = value
    if num_fmt:
        cell.number_format = num_fmt

def _next_input_filename(input_path, current_period, next_period):
    name = os.path.basename(input_path).replace("OwnerReport", "InputData")
    if current_period in name:
        return name.replace(current_period, next_period, 1)
    stem, ext = os.path.splitext(name)
    ext = ext or ".xlsx"
    replaced = re.sub(r"\d{4}-\d{2}$", next_period, stem)
    if replaced != stem:
        return f"{replaced}{ext}"
    return f"{stem}_{next_period}{ext}"

def _rollforward_snapshot(data, months):
    info = data["info"]
    cur = months[-1]
    days = _month_days(cur["year"], cur["month"])
    revpar = (cur["gross"] / days) if days else 0

    total_gross   = sum(m["gross"] for m in months)
    total_channel = sum(m["channel_fees"] for m in months)
    total_net     = sum(m["net_channel"] for m in months)
    total_tl      = sum(m["tl_share"] for m in months)
    total_owner   = sum(m["owner_share"] for m in months)
    total_exp     = sum(m["owner_exp"] for m in months)
    total_noi     = sum(m["net_owner_income"] for m in months)
    total_payouts = sum(m["payouts"] for m in months)

    existing_yield = _ratio_or_none(data.get("cumulative", {}).get("annualized_yield"))
    property_value = _property_purchase_price(info)
    annualized_yield = existing_yield
    if property_value and property_value > 0 and months:
        annualized_yield = (total_noi / property_value) * (12 / len(months))

    mgmt_start = info.get("mgmt_start_date")
    mgmt_start_str = (
        mgmt_start.strftime("%Y-%m-%d") if isinstance(mgmt_start, datetime)
        else str(mgmt_start or "")
    )

    return {
        "prior_period": {
            "gross_revenue":     _round_money(cur["gross"]),
            "channel_fees":      _round_money(cur["channel_fees"]),
            "net_after_channel": _round_money(cur["net_channel"]),
            "tl_share":          _round_money(cur["tl_share"]),
            "owner_share":       _round_money(cur["owner_share"]),
            "owner_expenses":    _round_money(cur["owner_exp"]),
            "net_owner_income":  _round_money(cur["net_owner_income"]),
            "cash_balance_end":  _round_money(cur["closing_bal"]),
            "occupancy_pct":     round(float(cur["occupancy"] or 0), 4),
            "adr":               _round_money(cur["adr"]),
            "revpar":            _round_money(revpar),
        },
        "cumulative": {
            "cumulative_gross_revenue":     _round_money(total_gross),
            "cumulative_channel_fees":      _round_money(total_channel),
            "cumulative_net_after_channel": _round_money(total_net),
            "cumulative_tl_share":          _round_money(total_tl),
            "cumulative_owner_share":       _round_money(total_owner),
            "cumulative_owner_expenses":    _round_money(total_exp),
            "cumulative_net_owner_income":  _round_money(total_noi),
            "total_owner_payouts":          _round_money(total_payouts),
            "annualized_yield": round(annualized_yield, 6) if annualized_yield is not None else None,
            "management_start_date": mgmt_start_str,
            "months_managed": len(months),
        },
        "cash_balance": {
            "opening_balance":      _round_money(cur["opening_bal"]),
            "total_owner_share":    _round_money(cur["owner_share"]),
            "total_owner_payin":    _round_money(cur["owner_payin"]),
            "total_owner_expenses": _round_money(-cur["owner_exp"]),
            "total_payouts":        _round_money(-cur["payouts"]),
            "closing_balance":      _round_money(cur["closing_bal"]),
        },
    }

def generate_next_input_template(input_path, output_path=None):
    data = read_input(input_path)
    info = data["info"]
    rpt_year, rpt_month = _parse_period(info)
    next_year, next_month = _next_period(rpt_year, rpt_month)
    current_period = _period_str(rpt_year, rpt_month)
    next_period = _period_str(next_year, next_month)

    mgmt_start = info.get("mgmt_start_date")
    if not isinstance(mgmt_start, datetime):
        raise ValueError("mgmt_start_date missing or invalid in Property_Info")

    months = compute_monthly(data, mgmt_start, rpt_year, rpt_month)
    if not months:
        raise ValueError("No monthly data computed — cannot roll forward next input template.")

    cur_month = months[-1]
    errors, warnings = validate(data, rpt_year, rpt_month, cur_month)
    if errors:
        raise ValueError("Validation errors; next input template not created:\n" + "\n".join(errors))

    wb = load_workbook(input_path)
    snapshot = _rollforward_snapshot(data, months)

    ws_info = wb["Property_Info"]
    _set_kv(ws_info, "period", next_period)
    _ensure_kv(
        ws_info,
        "asset_purchase_price",
        note="Стоимость покупки управляемой единицы (THB). Используется для annualized_yield.",
    )

    ws_prior = wb["Prior_Period"]
    for key, value in snapshot["prior_period"].items():
        num_fmt = "0.00%" if key == "occupancy_pct" else None
        _set_kv(ws_prior, key, value, num_fmt=num_fmt)

    ws_cumul = wb["Cumulative"]
    for key, value in snapshot["cumulative"].items():
        num_fmt = "0.00%" if key == "annualized_yield" else None
        _set_kv(ws_cumul, key, value, num_fmt=num_fmt)

    ws_cash = wb["Cash_Balance"]
    for key, value in snapshot["cash_balance"].items():
        _set_kv(ws_cash, key, value, num_fmt="#,##0.00")

    output_name = _next_input_filename(input_path, current_period, next_period)
    metadata = {
        "current_period": current_period,
        "next_period": next_period,
        "next_input_name": output_name,
        "opening_balance": snapshot["cash_balance"]["closing_balance"],
        "warnings": warnings,
    }

    if output_path:
        wb.save(output_path)
        wb.close()
        return metadata

    buf = io.BytesIO()
    wb.save(buf)
    wb.close()
    buf.seek(0)
    return buf.getvalue(), metadata


# ── MAIN ENTRY POINT ──────────────────────────────────────────────────────────
def generate_report(input_path, output_path=None):
    data = read_input(input_path)
    info = data["info"]
    rpt_year, rpt_month = _parse_period(info)

    mgmt_start = info.get("mgmt_start_date")
    if not isinstance(mgmt_start, datetime):
        raise ValueError("mgmt_start_date missing or invalid in Property_Info")

    months = compute_monthly(data, mgmt_start, rpt_year, rpt_month)
    if not months:
        raise ValueError("No monthly data computed — check period and management start date.")

    cur_month = months[-1]

    errors, warnings = validate(data, rpt_year, rpt_month, cur_month)
    if errors:
        raise ValueError("Validation errors:\n" + "\n".join(errors))

    wb = Workbook()
    wb.remove(wb.active)

    build_dashboard(wb, data, rpt_year, rpt_month, cur_month)
    build_pl(wb, data, rpt_year, rpt_month, cur_month)
    build_opex_passport(wb, data, rpt_year, rpt_month)
    build_12month(wb, data, rpt_year, rpt_month, months)
    build_ledger(wb, data, rpt_year, rpt_month, cur_month)
    build_dap(wb, data, rpt_year, rpt_month, months)

    tab_colors = {
        "Dashboard":          NAVY,
        "P&L Monthly":        TEAL,
        "OPEX Passport":      GOLD,
        "12-Month Summary":   NAVY,
        "Transaction Ledger": TEAL,
        "DAP Snapshot":       GOLD,
    }
    for sheet in wb.worksheets:
        sheet.sheet_properties.tabColor = tab_colors.get(sheet.title, NAVY)

    if output_path:
        wb.save(output_path)
        return warnings
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue(), warnings

def generate_report_bundle(input_path):
    """Return owner report bytes + next-month InputData bytes for Make.com."""
    report_bytes, warnings = generate_report(input_path)
    next_input_bytes, next_input_meta = generate_next_input_template(input_path)
    return {
        "report_bytes": report_bytes,
        "next_input_bytes": next_input_bytes,
        "warnings": warnings,
        "next_input": next_input_meta,
    }


# ── CLI ───────────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    parser = argparse.ArgumentParser(
        description="Generate TropicLook Owner Report (25/75 after OTA) and optionally roll forward next InputData."
    )
    parser.add_argument("input", help="Current TL_[CODE]_InputData_[YYYY-MM].xlsx file")
    parser.add_argument("output", help="Output TL_[CODE]_OwnerReport_[YYYY-MM].xlsx file")
    parser.add_argument(
        "--next-input",
        dest="next_input",
        help="Optional output path for prefilled next-month TL_[CODE]_InputData_[YYYY-MM].xlsx",
    )
    args = parser.parse_args()

    inp, out = args.input, args.output
    if not os.path.exists(inp):
        print(f"ERROR: Input file not found: {inp}"); sys.exit(1)

    try:
        warnings = generate_report(inp, out)
        print(f"OK Report generated ({ENGINE_VERSION}): {out}")
        for w in warnings:
            print(f"  WARN {w}")
        if args.next_input:
            meta = generate_next_input_template(inp, args.next_input)
            print(
                "OK Next input template generated: "
                f"{args.next_input} ({meta['current_period']} → {meta['next_period']}, "
                f"opening_balance={meta['opening_balance']:,.2f})"
            )
    except ValueError as e:
        print(f"ERROR: {e}")
        sys.exit(1)
